#!/usr/bin/env python3
"""Wipe-and-reload migration: GO Manager Data.xlsx -> Supabase Postgres.
Usage:
  python3 db/migrate_from_xlsx.py --dry-run     # parse + report only
  SUPABASE_DB_URL=postgresql://... python3 db/migrate_from_xlsx.py  # full load
Never touches Google Sheets or the live site."""
import argparse, json, os, re, sys, uuid
from collections import defaultdict
from datetime import datetime, timezone
from openpyxl import load_workbook

XLSX = os.path.join(os.path.dirname(__file__), '..', 'GO Manager Data.xlsx')
SET_KINDS = {'photocard', 'member', 'member-set'}
report = defaultdict(list)     # category -> [detail rows]

def parse_int(v):
    """Mirror of JS parseInt: takes the leading integer prefix of a string
    ('2abc' -> 2), while still handling numeric cells and float-strings
    ('2.0' -> 2) the way the sheet actually stores set_num/min_secure."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (TypeError, ValueError):
            return 0
    m = re.match(r'\s*[-+]?\d+', str(v))
    return int(m.group()) if m else 0

def build_sets_from_claims(claims, members):
    """Exact port of index.html buildSetsFromClaims placement logic (lines 5189-5245):
    fairness order by (created_at, claim_id); pass 1 groups OT claims by
    (username, declared set) into atomic fresh-numbered sets; pass 2 spills
    individual claims to the first non-OT set where their member slot is free.
    Returns {claim_id: set_no}. `members` is unused for placement (slots are a
    dict) but kept for signature parity / future validation."""
    ordered = sorted(claims, key=lambda c: (c.get('created_at') or '', str(c.get('claim_id'))))
    assignment = {}
    sets_by_num = {}   # set_no -> {member_name: claim_id}
    ot_set_nums = set()

    ot_groups = {}     # (username, declared) -> [claims], insertion-ordered
    for c in ordered:
        if c.get('assigned_vers') != 'OT':
            continue
        key = (c.get('username'), parse_int(c.get('set_num')) or 1)
        ot_groups.setdefault(key, []).append(c)
    for (_, declared), group in ot_groups.items():
        n = declared
        while n in sets_by_num:      # never reuse a number already owned by another set
            n += 1
        sets_by_num[n] = {}
        ot_set_nums.add(n)
        for c in group:
            sets_by_num[n][c.get('member_or_version')] = c['claim_id']
        for cid in sets_by_num[n].values():
            assignment[cid] = n

    def smallest_non_ot():
        n = 1
        while n in sets_by_num or n in ot_set_nums:
            n += 1
        return n

    for c in ordered:
        if c.get('assigned_vers') == 'OT':
            continue
        member = c.get('member_or_version')
        declared = parse_int(c.get('set_num')) or 1
        higher = sorted(x for x in sets_by_num if x > declared and x not in ot_set_nums)
        placed = False
        for n in [declared, *higher]:
            if n in ot_set_nums:
                continue
            slots = sets_by_num.setdefault(n, {})
            if member not in slots:
                slots[member] = c['claim_id']
                assignment[c['claim_id']] = n
                placed = True
                break
        if not placed:
            n = smallest_non_ot()
            sets_by_num.setdefault(n, {})[member] = c['claim_id']
            assignment[c['claim_id']] = n
    return assignment

def rows_as_dicts(ws):
    header = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, r))
        if any(v not in (None, '') for v in d.values()):
            out.append(d)
    return out

def to_ts(v, ctx=''):
    if v in (None, ''):
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s.replace('Z', '+00:00'))
    except ValueError:
        report['unparseable_dates'].append(f'{ctx}: {s!r}')
        return None

def norm_user(v):
    return re.sub(r'^@', '', str(v or '').strip())

def parse_members(val):
    """Mirror of index.html parseMembers: JSON array first, then newline/comma split."""
    if val in (None, ''):
        return []
    s = str(val).strip()
    if s in ('', '[]', '""'):
        return []
    try:
        parsed = json.loads(s)
        if isinstance(parsed, list):
            return [str(x).strip() for x in parsed if str(x).strip()]
    except (json.JSONDecodeError, ValueError):
        pass
    sep = '\n' if '\n' in s else ','
    return [x.strip() for x in s.split(sep) if x.strip()]

def parse_money(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0

def artist_from_name(name):
    return name.split(' - ')[0].strip() if name and ' - ' in name else None

def resolve_go_sheets(wb, gos, joiners):
    """Map go_id -> worksheet. Google's xlsx export truncates tab names to 31
    chars; colliding names get renamed 'SheetN'. Match uniquely-truncated names
    directly; disambiguate the rest by intersecting each sheet's sub_item_ids
    with each GO's claimed sub_item_ids from the joiners sheet."""
    si_sheets = {ws.title: ws for ws in wb.worksheets if ws['A1'].value == 'sub_item_id'}
    sheet_siids = {t: {r['sub_item_id'] for r in rows_as_dicts(ws) if r.get('sub_item_id')}
                   for t, ws in si_sheets.items()}
    claims_siids = defaultdict(set)
    for c in joiners:
        if c.get('go_id') and c.get('sub_item_id'):
            claims_siids[c['go_id']].add(c['sub_item_id'])

    by_trunc = defaultdict(list)
    for g in gos:
        by_trunc[('go_' + g['go_id'])[:31]].append(g['go_id'])

    mapping, taken = {}, set()
    ambiguous = []
    for trunc, ids in by_trunc.items():
        if len(ids) == 1 and trunc in si_sheets:
            mapping[ids[0]] = trunc
            taken.add(trunc)
        else:
            ambiguous.extend(ids)
    remaining_sheets = [t for t in si_sheets if t not in taken]
    for gid in list(ambiguous):
        want = claims_siids.get(gid, set())
        hits = [t for t in remaining_sheets if t not in taken and sheet_siids[t] & want]
        if len(hits) == 1:
            mapping[gid] = hits[0]
            taken.add(hits[0])
            ambiguous.remove(gid)
    # Elimination for claimless GOs: if exactly one GO and one sheet remain in a
    # truncation group, pair them.
    for gid in list(ambiguous):
        left = [t for t in remaining_sheets if t not in taken]
        if len([g for g in ambiguous]) == 1 and len(left) == 1:
            mapping[gid] = left[0]
            taken.add(left[0])
            ambiguous.remove(gid)
    for gid in ambiguous:
        report['unresolved_go_sheets'].append(gid)
    for t in si_sheets:
        if t not in taken:
            report['unmatched_subitem_sheets'].append(t)
    return {gid: si_sheets[t] for gid, t in mapping.items()}

def build_dataset(wb):
    """Everything the loader needs, as plain dicts keyed by legacy ids.
    Returns dict with keys: gos, sub_items, members, versions, sets, claims,
    payments, gc, shipping, shipping_items, listings, variants, shop_orders,
    store_orders. All FK references are legacy ids resolved to uuids at load."""
    gos = rows_as_dicts(wb['_gos'])
    all_joiner_rows = rows_as_dicts(wb['joiners'])
    joiners = [c for c in all_joiner_rows if c.get('claim_id')]
    n_dropped_joiners = len(all_joiner_rows) - len(joiners)
    if n_dropped_joiners:
        report['dropped_missing_id'].append(f'joiners: {n_dropped_joiners} rows missing claim_id')
    go_sheets = resolve_go_sheets(wb, gos, joiners)
    flags_closed = {(r['go_id'], r['sub_item_id']) for r in rows_as_dicts(wb['closed_subitems'])}
    flags_dead = {(r['go_id'], r['sub_item_id']): r['deadline'] for r in rows_as_dicts(wb['subitem_deadlines'])}
    flags_pay = {(r['go_id'], r['sub_item_id']): r['due_date'] for r in rows_as_dicts(wb['subitem_payment_due'])}
    secured = {(r['go_id'], r['sub_item_id'], parse_int(r['set_num']))
               for r in rows_as_dicts(wb['secured_sets'])}
    ds = {'gos': [], 'sub_items': [], 'members': [], 'versions': [], 'sets': [],
          'claims': [], 'payments': [], 'gc': [], 'shipping': [], 'shipping_items': [],
          'listings': [], 'variants': [], 'shop_orders': [], 'store_orders': []}

    claims_by_go = defaultdict(list)
    for c in joiners:
        claims_by_go[c.get('go_id')].append(c)

    known_go_ids = {g['go_id'] for g in gos}
    for c in joiners:
        if c.get('go_id') not in known_go_ids:
            report['orphan_claims_unknown_go'].append(c['claim_id'])

    for g in gos:
        gid = g['go_id']
        ds['gos'].append({
            'legacy_id': gid, 'name': g['name'], 'artist': artist_from_name(g['name']),
            'type': g.get('type') or 'photocard', 'status': g.get('status') or 'open',
            'deadline': to_ts(g.get('deadline'), f'go {gid} deadline'),
            'payment_deadline': to_ts(g.get('payment_deadline'), f'go {gid} pay_deadline'),
            'min_secure': parse_int(g.get('min_secure')) or 7,
            'created_at': to_ts(g.get('created_at'), f'go {gid} created'),
        })
        ws = go_sheets.get(gid)
        if ws is None:
            report['gos_without_subitem_sheet'].append(gid)
            continue
        si_rows = [r for r in rows_as_dicts(ws) if r.get('sub_item_id')]
        si_ids = {r['sub_item_id'] for r in si_rows}

        # Frontend orphan-remap (index.html:5014-5032): claims pointing at a
        # sub_item_id that no longer exists get remapped to the single
        # empty-members sub-item, if exactly one exists.
        orphans = [c for c in claims_by_go[gid] if c['sub_item_id'] not in si_ids]
        if orphans:
            empties = [r for r in si_rows if not parse_members(r.get('members'))]
            if len(empties) == 1:
                for c in orphans:
                    c['sub_item_id'] = empties[0]['sub_item_id']
                report['remapped_orphan_claims'].append(
                    f"{gid}: {len(orphans)} claims -> {empties[0]['sub_item_id']}")
            else:
                report['orphan_claims_no_target'].extend(c['claim_id'] for c in orphans)

        for pos, r in enumerate(si_rows):
            siid = r['sub_item_id']
            members = parse_members(r.get('members'))
            if not members:
                # Frontend fallback (index.html:5040-5044): reconstruct from claims.
                seen = {c['member_or_version'] for c in claims_by_go[gid]
                        if c['sub_item_id'] == siid and c.get('member_or_version')}
                members = sorted(seen)
                if members:
                    report['members_reconstructed_from_claims'].append(f'{gid}/{siid}')
            versions = parse_members(r.get('versions'))
            ms = parse_int(r.get('min_secure')) or 7
            kind = r.get('kind') or 'random'
            if kind not in ('photocard', 'member', 'member-set', 'single', 'versioned', 'random'):
                report['unknown_kinds'].append(f'{gid}/{siid}: {kind!r}')
                kind = 'random'
            ds['sub_items'].append({
                'legacy_id': siid, 'go': gid, 'name': r.get('name') or '',
                'kind': kind,
                'order_mode': 'batch' if ms < 0 else 'set',
                'batch_size': abs(ms) if ms < 0 else None,
                'price': parse_money(r.get('price')), 'ot_price': parse_money(r.get('ot_price')),
                'min_secure': ms if ms >= 0 else None,
                'image_url': r.get('image_url') or None, 'position': pos,
                'closed': (gid, siid) in flags_closed,
                'deadline': to_ts(flags_dead.get((gid, siid)), f'{siid} deadline'),
                'pay_due': to_ts(flags_pay.get((gid, siid)), f'{siid} pay_due'),
            })
            for i, m in enumerate(members):
                ds['members'].append({'si': (gid, siid), 'name': m, 'position': i})
            for i, v in enumerate(versions):
                ds['versions'].append({'si': (gid, siid), 'name': v, 'position': i})

            # Set replay for set-mechanic sub-items.
            si_claims = [c for c in claims_by_go[gid] if c['sub_item_id'] == siid]
            member_set = set(members)
            if kind in SET_KINDS and ms >= 0:
                placeable = [c for c in si_claims if c.get('member_or_version')]
                for c in placeable:
                    if c['member_or_version'] not in member_set and c.get('assigned_vers') != 'OT':
                        report['unknown_member_names'].append(
                            f"{c['claim_id']}: {c['member_or_version']!r} not in {gid}/{siid}")
                assignment = build_sets_from_claims(
                    [{'claim_id': c['claim_id'], 'username': norm_user(c.get('username')),
                      'member_or_version': c['member_or_version'],
                      'set_num': c.get('set_num'),
                      'created_at': str(c.get('created_at') or ''),
                      'assigned_vers': c.get('assigned_vers') or ''}
                     for c in placeable], members)
                for n in sorted(set(assignment.values())):
                    ds['sets'].append({'si': (gid, siid), 'set_no': n,
                                       'status': 'secured' if (gid, siid, n) in secured else 'open'})
                for c in si_claims:
                    c['_set_no'] = assignment.get(c['claim_id'])
                    if c.get('member_or_version') and c['claim_id'] not in assignment:
                        report['set_claims_left_unplaced'].append(c['claim_id'])

    known_si = {(s['go'], s['legacy_id']): s for s in ds['sub_items']}
    member_names = {(m['si'], m['name']) for m in ds['members']}
    version_names = {(v['si'], v['name']) for v in ds['versions']}
    for c in joiners:
        key = (c.get('go_id'), c.get('sub_item_id'))
        si = known_si.get(key)
        if si is None:
            continue    # already reported as orphan
        mv = c.get('member_or_version') or ''
        is_set_claim = si['kind'] in SET_KINDS and si['order_mode'] == 'set' and c.get('_set_no')
        member_ref = key if (mv and (key, mv) in member_names) else None
        version_ref = key if (mv and (key, mv) in version_names) else None
        ds['claims'].append({
            'legacy_id': c['claim_id'], 'si': key,
            'username': norm_user(c.get('username')),
            'email': c.get('email') or None,
            'set_no': c.get('_set_no') if is_set_claim else None,
            'member': mv if member_ref else None,
            'version': mv if (version_ref and not member_ref) else None,
            'is_ot': c.get('assigned_vers') == 'OT',
            'qty': max(parse_int(c.get('qty')), 1),
            'assigned_version': (c.get('assigned_vers') or None)
                                 if c.get('assigned_vers') != 'OT' else None,
            'status': c.get('claim_status') or 'pending',
            'payment_status': c.get('payment_status') or 'unpaid',
            'fulfillment': c.get('fulfillment') or 'Pending',
            'created_at': to_ts(c.get('created_at'), f"claim {c['claim_id']}"),
            'updated_at': to_ts(c.get('updated_at'), f"claim {c['claim_id']}"),
        })

    for idx, p in enumerate(rows_as_dicts(wb['payments'])):
        if not p.get('payment_id'):
            report['dropped_missing_id'].append(f'payments row {idx}: missing payment_id')
            continue
        ds['payments'].append({
            'legacy_id': p['payment_id'], 'username': norm_user(p.get('username')),
            'go': p.get('go_id') if p.get('go_id') in known_go_ids else None,
            'is_shop': p.get('go_id') == 'shop',
            'amount': parse_money(p.get('amount')), 'method': p.get('method'),
            'transaction_id': p.get('transaction_id') or None,
            'proof_url': p.get('proof_url') or None, 'email': p.get('email') or None,
            'status': p.get('status') or 'pending', 'note': p.get('note') or None,
            'created_at': to_ts(p.get('created_at'), f"payment {p['payment_id']}"),
        })
        if p.get('go_id') and p['go_id'] not in known_go_ids and p['go_id'] != 'shop':
            report['payments_unknown_go'].append(f"{p['payment_id']}: {p['go_id']}")

    for r in rows_as_dicts(wb['gc_added']):
        if r.get('go_id') in known_go_ids and r.get('username'):
            ds['gc'].append({'go': r['go_id'], 'username': norm_user(r['username'])})

    for idx, s in enumerate(rows_as_dicts(wb['shipping'])):
        if not s.get('request_id'):
            report['dropped_missing_id'].append(f'shipping row {idx}: missing request_id')
            continue
        ds['shipping'].append({
            'legacy_id': s['request_id'], 'username': norm_user(s.get('username')),
            'full_name': s.get('full_name'), 'address1': s.get('address1'),
            'address2': s.get('address2') or None, 'city': s.get('city'),
            'state': s.get('state'), 'postal': s.get('postal'), 'country': s.get('country'),
            'notes': s.get('notes') or None, 'email': s.get('email') or None,
            'ems_fee': parse_money(s['ems_fee']) if s.get('ems_fee') not in (None, '') else None,
            'dom_fee': parse_money(s['dom_fee']) if s.get('dom_fee') not in (None, '') else None,
            'total_fee': parse_money(s['total_fee']) if s.get('total_fee') not in (None, '') else None,
            'shipped': bool(s.get('shipped')),
            'created_at': to_ts(s.get('created_at'), f"shipping {s['request_id']}"),
        })
        try:
            items = json.loads(s.get('items') or '[]')
        except (json.JSONDecodeError, ValueError):
            items = []
            report['bad_shipping_items_json'].append(s['request_id'])
        for it in items:
            ds['shipping_items'].append({'request': s['request_id'], 'go': None,
                                         'description': it.get('label') or '', 'qty': 1})

    for idx, l in enumerate(rows_as_dicts(wb['listings'])):
        if not l.get('listing_id'):
            report['dropped_missing_id'].append(f'listings row {idx}: missing listing_id')
            continue
        try:
            variants = json.loads(l.get('variants') or 'null') or []
        except (json.JSONDecodeError, ValueError):
            variants = []
            report['bad_variants_json'].append(l['listing_id'])
        ds['listings'].append({
            'legacy_id': l['listing_id'], 'name': l.get('name') or '',
            'category': l.get('category'), 'price': parse_money(l.get('price')),
            'image_url': l.get('image_url') or None,
            'qty': parse_int(l.get('qty')) if not variants else None,
            'note': l.get('note') or None, 'status': l.get('status') or 'active',
            'created_at': to_ts(l.get('created_at'), f"listing {l['listing_id']}"),
        })
        for v in variants:
            ds['variants'].append({'listing': l['listing_id'], 'name': v.get('name'),
                                   'qty': parse_int(v.get('qty'))})

    known_listings = {l['legacy_id'] for l in ds['listings']}
    variant_names = {(v['listing'], v['name']) for v in ds['variants']}
    for idx, o in enumerate(rows_as_dicts(wb['shop_orders'])):
        if not o.get('order_id'):
            report['dropped_missing_id'].append(f'shop_orders row {idx}: missing order_id')
            continue
        if o.get('listing_id') not in known_listings:
            report['shop_orders_unknown_listing'].append(o['order_id'])
            continue
        ds['shop_orders'].append({
            'legacy_id': o['order_id'], 'listing': o['listing_id'],
            'variant': (o['listing_id'], o['variant'])
                       if o.get('variant') and (o['listing_id'], o['variant']) in variant_names
                       else None,
            'username': norm_user(o.get('username')), 'email': o.get('email') or None,
            'qty': max(parse_int(o.get('qty')), 1), 'unit_price': parse_money(o.get('unit_price')),
            'payment_status': o.get('payment_status') or 'unpaid',
            'fulfillment': o.get('fulfillment') or 'Ready',
            'created_at': to_ts(o.get('created_at'), f"shop {o['order_id']}"),
            'updated_at': to_ts(o.get('updated_at'), f"shop {o['order_id']}"),
        })

    for idx, o in enumerate(rows_as_dicts(wb['store_orders'])):
        if not o.get('order_id'):
            report['dropped_missing_id'].append(f'store_orders row {idx}: missing order_id')
            continue
        si_key = (o.get('go_id'), o.get('sub_item_id'))
        ds['store_orders'].append({
            'legacy_id': o['order_id'],
            'go': o.get('go_id') if o.get('go_id') in known_go_ids else None,
            'si': si_key if si_key in known_si else None,
            'store': o.get('store'), 'album_version': o.get('album_version') or None,
            'qty': max(parse_int(o.get('qty')), 1),
            'unit_cost': parse_money(o.get('unit_cost')) if o.get('unit_cost') not in (None, '') else None,
            'status': o.get('status'), 'notes': o.get('notes') or None,
            'created_at': to_ts(o.get('created_at'), f"store {o['order_id']}"),
            'updated_at': to_ts(o.get('updated_at'), f"store {o['order_id']}"),
        })
    return ds

def print_report(ds):
    print('=== intended row counts ===')
    for k in ds:
        print(f'  {k:16s} {len(ds[k])}')
    print('=== data-quality report ===')
    if not report:
        print('  clean — no issues found')
    for cat, items in sorted(report.items()):
        print(f'  {cat} ({len(items)}):')
        for it in items[:20]:
            print(f'    - {it}')
        if len(items) > 20:
            print(f'    … and {len(items) - 20} more')

def load(ds):
    raise SystemExit('load arrives in Task 8')

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    wb = load_workbook(XLSX)
    ds = build_dataset(wb)
    print_report(ds)
    if args.dry_run:
        return
    load(ds)   # Task 8

if __name__ == '__main__':
    main()
